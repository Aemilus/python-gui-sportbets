import itertools
from threading import Thread

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from core.bilet import Bilet


class GenereazaBilete:
    FOOTER_COLOR = PatternFill(start_color="afd9fa", end_color="afd9fa", fill_type="solid")
    NR_BILETE_VALIDE = 0

    def __init__(self, app):
        self.app = app
        self.thread = None
        self.combinatii = None
        self.bilete_valide = None
        self.tabel_bilete = None
        self.culori_semne = None
        self.wb = None
        self.ws = None

    def _init_tabel_bilete(self):
        """
        O lista de liste ce reprezinta un tabel.
        Scopul este ca ulterior:
        - pe primul rand se vor afla meciurile de pe pozitia 1 din toate biletele.
        - pe al doilea rand se vor afla meciurile de pe pozitia 2 din toate biletele.
        s.a.m.d. -> vezi filtreaza bilete.
        """
        self.tabel_bilete = list()
        for _ in range(self.app.core.nr_meciuri + 1):
            self.tabel_bilete.append(list())

    def _filtreaza_bilete(self):
        """
        Stocheaza biletele valide intr-un tabel.
        Biletele sunt distribuite pe coloane si spatiate cu o coloana goala intre.
        La baza biletului trecem numarul biletului.
        """
        for combinatie in self.combinatii:
            bilet = Bilet(self.app, combinatie)
            if bilet.is_valid():
                self.NR_BILETE_VALIDE += 1
                bilet.nr_bilet = self.NR_BILETE_VALIDE
                self.app.main_win.procesare_frame.label_progres_generare.config(text=str(self.NR_BILETE_VALIDE))
                for row in range(self.app.core.nr_meciuri):
                    self.tabel_bilete[row].append(bilet.combinatie[row].semn)
                    self.tabel_bilete[row].append("")
                self.tabel_bilete[self.app.core.nr_meciuri].append(f"Bilet {bilet.nr_bilet}")
                self.tabel_bilete[self.app.core.nr_meciuri].append("")

    def _dump_table_to_worksheet(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        for row in self.tabel_bilete:
            self.ws.append(row)

    def _color_worksheet(self):
        self.culori_semne = dict()
        for semn in self.app.core.semne:
            self.culori_semne[semn.semn] = semn.color
        # seteaza culoarea celulelor cu numarul biletului - footer-ul
        if self.tabel_bilete:
            for row in self.ws.iter_rows(
                    min_row=self.app.core.nr_meciuri+1,
                    max_row=self.app.core.nr_meciuri+1,
                    min_col=1,
                    max_col=len(self.tabel_bilete[0])
            ):
                for cell in row:
                    if cell.value:
                        cell.fill = self.FOOTER_COLOR
        # seteaza culoarea celulelor in functie de semn
        if self.tabel_bilete:
            for row in self.ws.iter_rows(
                    min_row=1, max_row=self.app.core.nr_meciuri, min_col=1, max_col=len(self.tabel_bilete[0])):
                for cell in row:
                    if str(cell.value):
                        cell.fill = self.culori_semne[str(cell.value)]

    def _save_to_excel(self):
        excel_path = str(self.app.path / "bilete.xlsx")
        self.wb.save(excel_path)
        self.app.core.excel = excel_path
        self.app.main_win.procesare_frame.label_progres_generare.config(text=f"Am generat {self.NR_BILETE_VALIDE} bilete")

    def _genereaza_bilete(self):
        # genereaza toate combinatiile
        self.combinatii = itertools.product(self.app.core.semne, repeat=self.app.core.nr_meciuri)
        self._init_tabel_bilete()
        self._filtreaza_bilete()
        self._dump_table_to_worksheet()
        self._color_worksheet()
        self._save_to_excel()

    def genereaza_bilete(self):
        if not self.thread:
            self.thread = Thread(target=self._genereaza_bilete, name="Thread-genereaza-bilete")
            self.thread.start()
